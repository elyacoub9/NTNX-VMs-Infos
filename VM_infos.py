import requests
import urllib3
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import getpass


urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)




pc_ip =input("your PC ip: ")
username = input("your PC user name: ")
password = getpass.getpass("your PC Password: ")
distination_file=input("distination file (.xlsx): ")

url = f'https://{pc_ip}:9440/api/nutanix/v3/vms/list'

payload = { 
    "kind": "vm",
    "sort_attribute": "vm_name",
    "length": 1000,
    "sort_order": "DESCENDING",
    "offset": 0}

headers = {'Content-Type': "application/json" }

auth = (username, password)

response = requests.post(url, json=payload, headers=headers, auth=auth,verify=False)

response.status_code == 200
data = response.json()

entities=data["entities"]
print(len(entities))
vms_infos=[]
 
for entity in entities:
    
    status=entity["status"]
    metadata=entity["metadata"]


    def vm_name_desc(Name_Or_Desc):
        if Name_Or_Desc=="Name":
            name=status["name"]
            return name

        elif Name_Or_Desc=="Desc":
            if "description" in status.keys():
                description=status["description"]
                return description
            else:
                return "no data "

    def vm_ram_cpu(RAM_or_CPU):
        if RAM_or_CPU=="CPU":
            vcpu=status["resources"]["num_sockets"]
            return vcpu
        elif RAM_or_CPU=="RAM":
            ram=status["resources"]["memory_size_mib"]
            ram=ram//1024
            return ram

    def vm_disks(nb_or_totalcapacity):
        if nb_or_totalcapacity=="nb":
            nb=0
            for disk in status["resources"]["disk_list"]:
                if disk["device_properties"]["device_type"]=="DISK":
                    nb+=1
                else:
                    continue
            return nb
        elif nb_or_totalcapacity=="totalcapacity":
            totalcapacity=0
            for disk in status["resources"]["disk_list"]:
                if disk["device_properties"]["device_type"]=="DISK":
                    disksize=int(disk["disk_size_mib"]/1024)
                    totalcapacity+=disksize

            return totalcapacity

    def vm_net(subnet_or_ip):
        subnets_names=[]
        ips=[]
        nic_list=status["resources"]["nic_list"]
        results=""
        if len (nic_list):
            for nic in nic_list:

                if subnet_or_ip=="subnet":
                    subnet_name=nic["subnet_reference"]["name"]
                    subnets_names.append(subnet_name)
                    result_list=subnets_names

                if subnet_or_ip=="ip":
                    if nic["ip_endpoint_list"] :
                        ip=nic["ip_endpoint_list"][0]["ip"]
                        ips.append(ip)
                        result_list=ips
                    else:
                        return "no data "
                
            results="\n".join(result_list)
            return(results.lstrip())
        else:
            return "no data "


    def vm_ngt(status_or_os):
        if "guest_tools" in status["resources"] :
            ntnx_gest_tools=status["resources"]["guest_tools"]["nutanix_guest_tools"]

            if status_or_os=="status":
                ngt_status=ntnx_gest_tools["ngt_state"]
                return ngt_status
            elif status_or_os=="os":

                os=ntnx_gest_tools["guest_os_version"]
                return os
        else:
            return "NGT not installed"

    def vm_powerstate_host(powerstate_or_host):
        power_state=status["resources"]["power_state"]
        if powerstate_or_host == "powerstate":
            return power_state

        if powerstate_or_host == "host":
            if power_state=="ON":
                host=status["resources"]["host_reference"]["name"]
                return host
            else:
                return f"{None} (the VM is OFF)"


    def vm_cluster():
        cluster=status["cluster_reference"]["name"]
        return cluster

    def vm_creation_time():

        creation_time=metadata["creation_time"]
        creation_time=datetime.fromisoformat(creation_time)
        creation_time=creation_time.strftime("%Y-%m-%d %H:%M")

        return creation_time

    def vm_categories():
        list_of_categories=[]
        categories_data=metadata["categories"]
        categories=""
        
        for key, value in categories_data.items():
            list_of_categories.append(f"{key}:{value}")
            
        categories="\n".join(list_of_categories)
        return(categories)
    

    vm= [vm_name_desc("Name"), vm_name_desc("Desc"),
           vm_ram_cpu("RAM"), vm_ram_cpu("CPU"),
             vm_disks("nb"), vm_disks("totalcapacity"),
                str(vm_net("subnet")), str(vm_net("ip")),
                  vm_ngt("status"), vm_ngt("os"),
                    vm_powerstate_host("powerstate"), vm_powerstate_host("host"),
                    vm_cluster(), vm_creation_time(),
                    str(vm_categories()) ]
    

       
    vms_infos.append(vm)


#link VMs with its efficiency status

url2 = f'https://{pc_ip}:9440/api/nutanix/v3/groups'

payload2 = {
    
    "entity_type": "mh_vm",
    "sort_attribute": "vm_name",
    "group_member_attributes": [
        {
            "attribute": "vm_name"
        },
        {
            "attribute": "capacity.vm_efficiency_status"
        }
    ]
}


response2 = requests.post(url2, json=payload2, headers=headers, auth=auth,verify=False)

data2 = response2.json()
#print((data["group_results"][0]["entity_results"][1]["data"][0]))
#print("---------")
#print((data["group_results"][0]["entity_results"][1]["data"][1]

def vm_efficiency(my_vm):
    vms_efficiency={}
    entity_results=data2["group_results"][0]["entity_results"]
    for entitie2 in entity_results:

        vm_name=entitie2["data"][0]["values"][0]["values"]
        efficiency=entitie2["data"][1]["values"][0]["values"]

        vms_efficiency["".join(vm_name)]="".join(efficiency)
    return (vms_efficiency[my_vm])

vmss=vms_infos
for i in range(len(vms_infos)):
    
    vm_name=vms_infos[i][0]
    vms_infos[i].append(str(vm_efficiency(vm_name)))

    

wb = Workbook()
ws = wb.active
ws.append(["VM name", "description", "Memory(Gib)", "vCPU", "number of disks", "total storage(Gib)", "Subnets", "IP address", "NGT Status", "OS", "Power State", "Host", "Cluster", "Creation Time", "categories","vm efficiency"])

ft = Font(bold=True) # <---thanks to Chat-GPT
for row in ws["A1:P1"]:
    for cell in row:
        cell.font = ft
        

for row in vms_infos:
    ws.append(row)
    
    for row in ws.iter_rows():# <---thanks to Chat-GPT
        for cell in row:
            alignment = Alignment(vertical='center')
            cell.alignment = alignment

for col in ws.columns:    # <---thanks to Chat-GPT
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        if len(str(cell.value)) > max_length:
            max_length = len(str(cell.value))
    adjusted_width = max_length
    ws.column_dimensions[column].width = adjusted_width 

wb.save(distination_file)



